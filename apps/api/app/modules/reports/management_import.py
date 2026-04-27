from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from apps.api.app.db.models import (
    ManagementReportImport,
    ManagementReportMetric,
    ManagementReportRow,
    OrganizationUnit,
)

PARSER_VERSION = "management_report_v1"

DEPARTMENT_RE = re.compile(r"^(?P<code>\d{4})\s+(?P<name>.+)$")


@dataclass(frozen=True)
class ManagementReportImportSummary:
    import_id: str
    file_name: str
    checksum: str
    sheet_count: int
    raw_row_count: int
    metric_count: int
    organization_unit_count: int


def _checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def _clean_label(value: Any) -> str | None:
    serialized = _serialize_cell(value)
    if serialized is None:
        return None
    cleaned = str(serialized).strip()
    return cleaned or None


def _is_empty_row(values: list[Any]) -> bool:
    return all(_serialize_cell(value) is None for value in values)


def _to_decimal(value: Any) -> Decimal | None:
    serialized = _serialize_cell(value)
    if serialized is None or isinstance(serialized, bool):
        return None
    if isinstance(serialized, int | float):
        return Decimal(str(serialized))
    text = str(serialized).replace("\xa0", "").replace(" ", "").replace("−", "-")
    text = text.replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _is_total(label: str) -> bool:
    return label.strip().lower() == "итого"


def _is_demand_shortage(label: str) -> bool:
    return label.strip().lower() == "спрос"


def _department_parts(label: str) -> tuple[str | None, str, str]:
    match = DEPARTMENT_RE.match(label.strip())
    if not match:
        return None, label.strip(), label.strip()
    code = match.group("code")
    name = match.group("name").strip()
    return code, name, label.strip()


def _dimension_for_label(default_type: str, label: str) -> tuple[str, str | None, str]:
    if _is_total(label):
        return "total", None, "Итого"
    if _is_demand_shortage(label):
        return "demand_shortage", None, "СПРОС"
    if default_type == "department":
        code, name, _ = _department_parts(label)
        return default_type, code, name
    return default_type, None, label.strip()


def _upsert_department(
    db: Session,
    *,
    report: ManagementReportImport,
    sheet_name: str,
    row_index: int,
    label: str,
) -> OrganizationUnit | None:
    code, name, display_name = _department_parts(label)
    if code is None or _is_total(label) or _is_demand_shortage(label):
        return None
    for pending in db.new:
        if (
            isinstance(pending, OrganizationUnit)
            and pending.unit_type == "department"
            and pending.code == code
        ):
            pending.name = name
            pending.display_name = display_name
            pending.active = True
            pending.source_import_id = report.id
            pending.metadata_payload = {"source_sheet": sheet_name, "source_row_index": row_index}
            return pending
    unit = db.scalar(
        select(OrganizationUnit).where(
            OrganizationUnit.unit_type == "department",
            OrganizationUnit.code == code,
        )
    )
    metadata = {"source_sheet": sheet_name, "source_row_index": row_index}
    if unit is None:
        unit = OrganizationUnit(
            unit_type="department",
            code=code,
            name=name,
            display_name=display_name,
            active=True,
            source_import_id=report.id,
            metadata_payload=metadata,
        )
        db.add(unit)
        return unit
    unit.name = name
    unit.display_name = display_name
    unit.active = True
    unit.source_import_id = report.id
    unit.metadata_payload = metadata
    return unit


def _add_metric(
    db: Session,
    *,
    report: ManagementReportImport,
    source_row: ManagementReportRow,
    sheet_name: str,
    row_index: int,
    row_values: list[Any],
    dimension_type: str,
    dimension_code: str | None,
    dimension_name: str,
    metric_name: str,
    metric_year: int | None,
    metric_unit: str,
    raw_value: Any,
    source_column_index: int,
) -> int:
    value = _to_decimal(raw_value)
    if value is None:
        return 0
    metric = ManagementReportMetric(
        import_id=report.id,
        source_row_id=source_row.id,
        sheet_name=sheet_name,
        row_index=row_index,
        dimension_type=dimension_type,
        dimension_code=dimension_code,
        dimension_name=dimension_name,
        metric_name=metric_name,
        metric_year=metric_year,
        metric_value=value,
        metric_unit=metric_unit,
        raw_payload={
            "parser_version": PARSER_VERSION,
            "source_column_index": source_column_index + 1,
            "raw_value": _serialize_cell(raw_value),
            "raw_row": [_serialize_cell(value) for value in row_values],
        },
    )
    db.add(metric)
    source_row.parsed_metric_count += 1
    return 1


def _parse_year_profitability_sheet(
    db: Session,
    *,
    report: ManagementReportImport,
    sheet_name: str,
    rows: list[tuple[int, list[Any]]],
    row_objects: dict[tuple[str, int], ManagementReportRow],
    dimension_type: str,
    start_row: int = 3,
) -> int:
    count = 0
    metric_specs = (
        (1, "revenue", 2024, "rub"),
        (2, "profitability_pct", 2024, "pct"),
        (3, "revenue", 2025, "rub"),
        (4, "profitability_pct", 2025, "pct"),
        (5, "growth_rate", None, "ratio"),
    )
    for row_index, values in rows:
        if row_index < start_row:
            continue
        label = _clean_label(values[0] if values else None)
        if not label:
            continue
        row_dimension_type, dimension_code, dimension_name = _dimension_for_label(
            dimension_type, label
        )
        if row_dimension_type == "department":
            _upsert_department(
                db, report=report, sheet_name=sheet_name, row_index=row_index, label=label
            )
        source_row = row_objects[(sheet_name, row_index)]
        for column_index, metric_name, year, unit in metric_specs:
            if column_index >= len(values):
                continue
            actual_metric_name = (
                "demand_shortage"
                if row_dimension_type == "demand_shortage" and metric_name == "revenue"
                else metric_name
            )
            count += _add_metric(
                db,
                report=report,
                source_row=source_row,
                sheet_name=sheet_name,
                row_index=row_index,
                row_values=values,
                dimension_type=row_dimension_type,
                dimension_code=dimension_code,
                dimension_name=dimension_name,
                metric_name=actual_metric_name,
                metric_year=year,
                metric_unit=unit,
                raw_value=values[column_index],
                source_column_index=column_index,
            )
    return count


def _parse_own_production_sheet(
    db: Session,
    *,
    report: ManagementReportImport,
    sheet_name: str,
    rows: list[tuple[int, list[Any]]],
    row_objects: dict[tuple[str, int], ManagementReportRow],
) -> int:
    count = 0
    section = "own_production_product_group"
    for row_index, values in rows:
        label = _clean_label(values[0] if values else None)
        if not label:
            continue
        if label.lower() == "подразделение":
            section = "own_production_department"
            continue
        if row_index <= 2 or label.lower() == "товарная группа":
            continue

        dimension_type, dimension_code, dimension_name = _dimension_for_label(
            "department" if section == "own_production_department" else section,
            label,
        )
        if dimension_type == "department":
            _upsert_department(
                db, report=report, sheet_name=sheet_name, row_index=row_index, label=label
            )

        source_row = row_objects[(sheet_name, row_index)]
        specs = (
            (1, "revenue", 2024, "rub"),
            (2, "revenue", 2025, "rub"),
            (3, "growth_rate", None, "ratio"),
        )
        for column_index, metric_name, year, unit in specs:
            if column_index >= len(values):
                continue
            count += _add_metric(
                db,
                report=report,
                source_row=source_row,
                sheet_name=sheet_name,
                row_index=row_index,
                row_values=values,
                dimension_type=dimension_type,
                dimension_code=dimension_code,
                dimension_name=dimension_name,
                metric_name=metric_name,
                metric_year=year,
                metric_unit=unit,
                raw_value=values[column_index],
                source_column_index=column_index,
            )
    return count


def _parse_network_sheet(
    db: Session,
    *,
    report: ManagementReportImport,
    sheet_name: str,
    rows: list[tuple[int, list[Any]]],
    row_objects: dict[tuple[str, int], ManagementReportRow],
) -> int:
    count = 0
    specs = (
        (1, "revenue", 2024, "rub"),
        (2, "profitability_pct", 2024, "pct"),
        (3, "revenue", 2025, "rub"),
        (4, "profitability_pct", 2025, "pct"),
        (5, "distribution_share", 2024, "ratio"),
        (6, "distribution_share", 2025, "ratio"),
    )
    for row_index, values in rows:
        if row_index < 3:
            continue
        label = _clean_label(values[0] if values else None)
        if not label:
            continue
        dimension_type, dimension_code, dimension_name = _dimension_for_label("network", label)
        source_row = row_objects[(sheet_name, row_index)]
        for column_index, metric_name, year, unit in specs:
            if column_index >= len(values):
                continue
            count += _add_metric(
                db,
                report=report,
                source_row=source_row,
                sheet_name=sheet_name,
                row_index=row_index,
                row_values=values,
                dimension_type=dimension_type,
                dimension_code=dimension_code,
                dimension_name=dimension_name,
                metric_name=metric_name,
                metric_year=year,
                metric_unit=unit,
                raw_value=values[column_index],
                source_column_index=column_index,
            )
    return count


def _parse_pdz_sheet(
    db: Session,
    *,
    report: ManagementReportImport,
    sheet_name: str,
    rows: list[tuple[int, list[Any]]],
    row_objects: dict[tuple[str, int], ManagementReportRow],
) -> int:
    count = 0
    specs = (
        (1, "total_receivables", 2024, "rub"),
        (2, "overdue_receivables", 2024, "rub"),
        (3, "total_receivables", 2025, "rub"),
        (4, "overdue_receivables", 2025, "rub"),
        (5, "overdue_receivables_reduction_rate", None, "ratio"),
    )
    for row_index, values in rows:
        if row_index < 3:
            continue
        label = _clean_label(values[0] if values else None)
        if not label:
            continue
        dimension_type, dimension_code, dimension_name = _dimension_for_label("counterparty", label)
        if dimension_code is None:
            department_type, department_code, department_name = _dimension_for_label(
                "department", label
            )
            if department_code is not None:
                dimension_type = department_type
                dimension_code = department_code
                dimension_name = department_name
                _upsert_department(
                    db, report=report, sheet_name=sheet_name, row_index=row_index, label=label
                )
        source_row = row_objects[(sheet_name, row_index)]
        for column_index, metric_name, year, unit in specs:
            if column_index >= len(values):
                continue
            count += _add_metric(
                db,
                report=report,
                source_row=source_row,
                sheet_name=sheet_name,
                row_index=row_index,
                row_values=values,
                dimension_type=dimension_type,
                dimension_code=dimension_code,
                dimension_name=dimension_name,
                metric_name=metric_name,
                metric_year=year,
                metric_unit=unit,
                raw_value=values[column_index],
                source_column_index=column_index,
            )
    return count


def _parse_simple_three_metric_sheet(
    db: Session,
    *,
    report: ManagementReportImport,
    sheet_name: str,
    rows: list[tuple[int, list[Any]]],
    row_objects: dict[tuple[str, int], ManagementReportRow],
    dimension_type: str,
    start_row: int,
) -> int:
    count = 0
    specs = (
        (1, "revenue", 2024, "rub"),
        (2, "revenue", 2025, "rub"),
        (3, "growth_rate", None, "ratio"),
    )
    for row_index, values in rows:
        if row_index < start_row:
            continue
        label = _clean_label(values[0] if values else None)
        if not label:
            continue
        row_dimension_type, dimension_code, dimension_name = _dimension_for_label(
            dimension_type, label
        )
        source_row = row_objects[(sheet_name, row_index)]
        for column_index, metric_name, year, unit in specs:
            if column_index >= len(values):
                continue
            count += _add_metric(
                db,
                report=report,
                source_row=source_row,
                sheet_name=sheet_name,
                row_index=row_index,
                row_values=values,
                dimension_type=row_dimension_type,
                dimension_code=dimension_code,
                dimension_name=dimension_name,
                metric_name=metric_name,
                metric_year=year,
                metric_unit=unit,
                raw_value=values[column_index],
                source_column_index=column_index,
            )
    return count


def _parse_demand_shortage_sheet(
    db: Session,
    *,
    report: ManagementReportImport,
    sheet_name: str,
    rows: list[tuple[int, list[Any]]],
    row_objects: dict[tuple[str, int], ManagementReportRow],
) -> int:
    count = 0
    for row_index, values in rows:
        if row_index < 2:
            continue
        source_row = row_objects[(sheet_name, row_index)]
        for column_index, year in ((0, 2024), (1, 2025)):
            if column_index >= len(values):
                continue
            count += _add_metric(
                db,
                report=report,
                source_row=source_row,
                sheet_name=sheet_name,
                row_index=row_index,
                row_values=values,
                dimension_type="demand_shortage",
                dimension_code=None,
                dimension_name="СПРОС (недопоставки)",
                metric_name="demand_shortage",
                metric_year=year,
                metric_unit="rub",
                raw_value=values[column_index],
                source_column_index=column_index,
            )
    return count


def _is_header_row(sheet_name: str, row_index: int, values: list[Any]) -> bool:
    first = (_clean_label(values[0]) or "").lower() if values else ""
    if row_index <= 2:
        return True
    return first in {"товарная группа", "подразделение", "контрагент", "регион", "тз"}


def import_management_report_workbook(
    db: Session,
    file_path: Path,
    *,
    report_year: int | None = 2025,
    imported_by_id: str | None = None,
    upload_file_id: str | None = None,
    source_path: str | None = None,
    file_name: str | None = None,
    commit: bool = True,
) -> ManagementReportImportSummary:
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    checksum = _checksum(file_path)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    stored_file_name = file_name or file_path.name

    report = db.scalar(
        select(ManagementReportImport).where(ManagementReportImport.checksum == checksum)
    )
    if report is None:
        report = ManagementReportImport(
            file_name=stored_file_name,
            source_path=source_path or str(file_path),
            checksum=checksum,
            report_year=report_year,
            imported_by_id=imported_by_id,
            upload_file_id=upload_file_id,
            metadata_payload={},
        )
        db.add(report)
        db.flush()
    else:
        db.execute(
            delete(ManagementReportMetric).where(ManagementReportMetric.import_id == report.id)
        )
        db.execute(delete(ManagementReportRow).where(ManagementReportRow.import_id == report.id))
        db.execute(delete(OrganizationUnit).where(OrganizationUnit.source_import_id == report.id))
        report.file_name = stored_file_name
        report.source_path = source_path or str(file_path)
        report.report_year = report_year
        report.imported_by_id = imported_by_id
        report.upload_file_id = upload_file_id
        db.flush()

    sheet_rows: dict[str, list[tuple[int, list[Any]]]] = {}
    row_objects: dict[tuple[str, int], ManagementReportRow] = {}
    raw_row_count = 0

    for worksheet in workbook.worksheets:
        rows: list[tuple[int, list[Any]]] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = list(row)
            if _is_empty_row(values):
                continue
            raw_values = [_serialize_cell(value) for value in values]
            report_row = ManagementReportRow(
                import_id=report.id,
                sheet_name=worksheet.title,
                row_index=row_index,
                is_header=_is_header_row(worksheet.title, row_index, values),
                raw_values=raw_values,
            )
            db.add(report_row)
            row_objects[(worksheet.title, row_index)] = report_row
            rows.append((row_index, values))
            raw_row_count += 1
        sheet_rows[worksheet.title] = rows

    db.flush()

    metric_count = 0
    metric_count += _parse_year_profitability_sheet(
        db,
        report=report,
        sheet_name="Товарная группа",
        rows=sheet_rows.get("Товарная группа", []),
        row_objects=row_objects,
        dimension_type="product_group",
    )
    metric_count += _parse_simple_three_metric_sheet(
        db,
        report=report,
        sheet_name="ТЗ",
        rows=sheet_rows.get("ТЗ", []),
        row_objects=row_objects,
        dimension_type="technical_zone",
        start_row=3,
    )
    metric_count += _parse_own_production_sheet(
        db,
        report=report,
        sheet_name="Продажи Собственного Производст",
        rows=sheet_rows.get("Продажи Собственного Производст", []),
        row_objects=row_objects,
    )
    metric_count += _parse_year_profitability_sheet(
        db,
        report=report,
        sheet_name="Подразделение",
        rows=sheet_rows.get("Подразделение", []),
        row_objects=row_objects,
        dimension_type="department",
    )
    metric_count += _parse_year_profitability_sheet(
        db,
        report=report,
        sheet_name="Регионы ОПТ",
        rows=sheet_rows.get("Регионы ОПТ", []),
        row_objects=row_objects,
        dimension_type="region_opt",
    )
    metric_count += _parse_network_sheet(
        db,
        report=report,
        sheet_name="Сети - разбивка по сетям",
        rows=sheet_rows.get("Сети - разбивка по сетям", []),
        row_objects=row_objects,
    )
    metric_count += _parse_pdz_sheet(
        db,
        report=report,
        sheet_name="ПДЗ",
        rows=sheet_rows.get("ПДЗ", []),
        row_objects=row_objects,
    )
    metric_count += _parse_demand_shortage_sheet(
        db,
        report=report,
        sheet_name="СПРОС (недопоставки)",
        rows=sheet_rows.get("СПРОС (недопоставки)", []),
        row_objects=row_objects,
    )

    report.sheet_count = len(workbook.sheetnames)
    report.raw_row_count = raw_row_count
    report.metric_count = metric_count
    report.metadata_payload = {
        "parser_version": PARSER_VERSION,
        "sheet_names": list(workbook.sheetnames),
        "source_file_size_bytes": file_path.stat().st_size,
        "upload_file_id": upload_file_id,
    }

    if commit:
        db.commit()
    else:
        db.flush()
    organization_unit_count = len(
        db.scalars(
            select(OrganizationUnit.id).where(OrganizationUnit.source_import_id == report.id)
        ).all()
    )
    db.refresh(report)
    return ManagementReportImportSummary(
        import_id=report.id,
        file_name=report.file_name,
        checksum=report.checksum,
        sheet_count=report.sheet_count,
        raw_row_count=report.raw_row_count,
        metric_count=report.metric_count,
        organization_unit_count=organization_unit_count,
    )
