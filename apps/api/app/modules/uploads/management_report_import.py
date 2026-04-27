from __future__ import annotations

import re
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from apps.api.app.common.utils import utc_now
from apps.api.app.core.errors import DomainError
from apps.api.app.db.models import QualityIssue, UploadedRowIssue
from apps.api.app.db.models import UploadFile as UploadFileModel
from apps.api.app.modules.reports.management_import import import_management_report_workbook
from apps.api.app.modules.uploads.storage import ObjectStorage

MANAGEMENT_REPORT_SHEETS = {
    "Товарная группа",
    "ТЗ",
    "Продажи Собственного Производст",
    "Подразделение",
    "Регионы ОПТ",
    "Сети - разбивка по сетям",
    "ПДЗ",
    "СПРОС (недопоставки)",
}


def _infer_report_year(file_name: str) -> int | None:
    match = re.search(r"\b(20\d{2})\b", file_name)
    return int(match.group(1)) if match else None


def is_management_report_payload(payload: bytes, file_name: str) -> bool:
    suffix = Path(file_name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        return False
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception:
        return False
    sheet_names = set(workbook.sheetnames)
    matched = sheet_names.intersection(MANAGEMENT_REPORT_SHEETS)
    return (
        "Товарная группа" in matched
        and ("Подразделение" in matched or "Сети - разбивка по сетям" in matched)
        and len(matched) >= 3
    )


def _write_temp_workbook(payload: bytes, file_name: str) -> Path:
    suffix = Path(file_name).suffix or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as handle:
        handle.write(payload)
        return Path(handle.name)


def import_management_report_from_upload(
    db: Session,
    storage: ObjectStorage,
    upload_file_id: str,
    *,
    commit: bool = True,
) -> None:
    upload_file = db.get(UploadFileModel, upload_file_id)
    if upload_file is None:
        raise DomainError(code="upload_file_not_found", message="Файл загрузки не найден")
    batch = upload_file.batch
    if batch is None:
        raise DomainError(code="upload_batch_not_found", message="Пакет загрузки не найден")

    payload = storage.load_upload_bytes(upload_file.storage_key)
    if not is_management_report_payload(payload, upload_file.file_name):
        raise DomainError(
            code="not_management_report",
            message="Файл не похож на управленческий отчёт MAGAMAX",
        )

    temp_path = _write_temp_workbook(payload, upload_file.file_name)
    try:
        summary = import_management_report_workbook(
            db,
            temp_path,
            report_year=_infer_report_year(upload_file.file_name),
            imported_by_id=batch.uploaded_by_id,
            upload_file_id=upload_file.id,
            source_path=f"object://{upload_file.storage_key}",
            file_name=upload_file.file_name,
            commit=False,
        )
    finally:
        temp_path.unlink(missing_ok=True)

    db.execute(delete(UploadedRowIssue).where(UploadedRowIssue.batch_id == batch.id))
    db.execute(
        delete(QualityIssue).where(
            QualityIssue.batch_id == batch.id,
            QualityIssue.file_id == upload_file.id,
        )
    )

    upload_file.source_type = "management_report"
    batch.source_type = "management_report"
    batch.detected_source_type = "management_report"
    batch.total_rows = summary.raw_row_count
    batch.valid_rows = summary.raw_row_count
    batch.applied_rows = summary.raw_row_count
    batch.failed_rows = 0
    batch.warning_count = 0
    batch.issue_count = 0
    batch.status = "applied"
    batch.mapping_template_id = None
    batch.mapping_payload = {
        "suggestions": [],
        "active_mapping": {},
        "required_fields": [],
        "management_report_adapter": {
            "parser_version": "management_report_v1",
            "import_id": summary.import_id,
            "sheet_count": summary.sheet_count,
            "raw_row_count": summary.raw_row_count,
            "metric_count": summary.metric_count,
            "organization_unit_count": summary.organization_unit_count,
        },
        "source_detection": {
            **dict((batch.mapping_payload or {}).get("source_detection") or {}),
            "requires_confirmation": False,
            "confirmed": True,
            "detected_source_type": "management_report",
            "selected_source_type": "management_report",
            "candidates": [
                {
                    "source_type": "management_report",
                    "confidence": 0.99,
                    "matched_fields": ["workbook_sheets", "management_metrics"],
                }
            ],
        },
    }
    batch.preview_payload = {
        **dict(batch.preview_payload or {}),
        "management_report": {
            "import_id": summary.import_id,
            "sheet_count": summary.sheet_count,
            "raw_row_count": summary.raw_row_count,
            "metric_count": summary.metric_count,
        },
    }
    batch.validation_payload = {
        "validated_at": utc_now().isoformat(),
        "valid_rows": summary.raw_row_count,
        "failed_rows": 0,
        "warning_count": 0,
        "has_blocking_issues": False,
        "issue_counts": {"info": 0, "warning": 0, "error": 0, "critical": 0, "total": 0},
        "source_type": "management_report",
        "import_id": summary.import_id,
        "metric_count": summary.metric_count,
        "sheet_count": summary.sheet_count,
    }
    status_payload = dict(batch.status_payload or {})
    history = list(status_payload.get("history", []))
    history.append(
        {
            "status": "applied",
            "at": utc_now().isoformat(),
            "message": "Управленческий отчёт импортирован и связан с Upload Center",
            "error": None,
        }
    )
    batch.status_payload = {
        **status_payload,
        "history": history[-25:],
        "current_status_at": utc_now().isoformat(),
        "last_error": None,
    }

    if commit:
        db.commit()
    else:
        db.flush()
