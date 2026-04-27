from __future__ import annotations

import hashlib
from dataclasses import dataclass
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from sqlalchemy import delete, func, select, update
from sqlalchemy.orm import Session

from apps.api.app.common.utils import utc_now
from apps.api.app.db.models import (
    Category,
    InboundDelivery,
    Product,
    QualityIssue,
    ReserveRow,
    SalesFact,
    Sku,
    SkuCost,
    StockSnapshot,
    UploadedRowIssue,
    UploadFile,
)
from apps.api.app.modules.catalog.brand import resolve_brand
from apps.api.app.modules.uploads.storage import ObjectStorage


@dataclass(frozen=True)
class AssortmentImportSummary:
    upload_file_id: str
    file_name: str
    raw_rows: int
    category_rows: int
    product_rows: int
    linked_sku_rows: int
    deleted_bad_sku_rows: int
    skipped_rows: int


ARTICLE_HEADERS = ("артикул", "article", "sku", "sku_code")
PRODUCT_HEADERS = ("номенклатура", "наименование", "товар", "product_name")
NON_ASSORTMENT_HEADERS = (
    "себестоимость",
    "себесстоимость",
    "cost",
    "доступно",
    "остаток",
    "stock",
    "quantity",
    "qty",
)


def _normalize_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip().lower().replace("ё", "е")


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def _find_column(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {_normalize_text(header): header for header in headers}
    for candidate in candidates:
        candidate_normalized = _normalize_text(candidate)
        for header_normalized, original in normalized.items():
            if candidate_normalized in header_normalized:
                return original
    return None


def is_assortment_frame(frame: pd.DataFrame, file_name: str | None = None) -> bool:
    headers = [str(column) for column in frame.columns]
    if not _find_column(headers, ARTICLE_HEADERS) or not _find_column(headers, PRODUCT_HEADERS):
        return False
    header_blob = " ".join(_normalize_text(header) for header in headers)
    if any(candidate in header_blob for candidate in NON_ASSORTMENT_HEADERS):
        return False
    file_blob = _normalize_text(file_name or "")
    if "ассортимент" in file_blob or "assort" in file_blob:
        return True
    article_column = _find_column(headers, ARTICLE_HEADERS)
    product_column = _find_column(headers, PRODUCT_HEADERS)
    if article_column is None or product_column is None:
        return False
    labels = [_normalize_text(value) for value in frame[article_column].head(120).tolist()]
    product_values = [_clean_text(value) for value in frame[product_column].head(120).tolist()]
    category_like = sum(
        1
        for value, product in zip(labels, product_values, strict=False)
        if value and not product and ("магамакс" in value or value[:2].isdigit())
    )
    product_like = sum(1 for value in product_values if value)
    return category_like >= 3 and product_like >= 1


def _cell_text(cell: Cell) -> str | None:
    return _clean_text(cell.value)


def _find_header_columns(ws: Any) -> tuple[int, int, int]:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        values = [_normalize_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, 15)]
        article_col = None
        product_col = None
        for col_idx, value in enumerate(values, start=1):
            if article_col is None and any(item in value for item in ARTICLE_HEADERS):
                article_col = col_idx
            if product_col is None and any(item in value for item in PRODUCT_HEADERS):
                product_col = col_idx
        if article_col is not None and product_col is not None:
            return row_idx, article_col, product_col
    raise ValueError("Assortment file must contain article and product columns")


def _category_code(path: str) -> str:
    digest = hashlib.sha1(path.encode("utf-8")).hexdigest()[:10]
    readable = (
        path.lower()
        .replace("ё", "е")
        .replace(" / ", "_")
        .replace(" ", "_")
        .replace(".", "_")
        .replace("-", "_")
    )
    readable = "".join(char for char in readable if char.isalnum() or char == "_")
    readable = "_".join(item for item in readable.split("_") if item)
    prefix = readable[:58] or "category"
    return f"{prefix}_{digest}"[:80]


def _get_or_create_category(
    db: Session,
    *,
    name: str,
    parent: Category | None,
    level: int,
    path: str,
) -> Category:
    existing = db.scalars(select(Category).where(Category.path == path)).first()
    if existing is not None:
        existing.name = name
        existing.parent_id = parent.id if parent else None
        existing.level = level
        return existing
    category = Category(
        code=_category_code(path),
        name=name,
        parent_id=parent.id if parent else None,
        level=level,
        path=path,
    )
    db.add(category)
    db.flush()
    return category


def _get_or_create_product(db: Session, name: str, brand: str) -> Product:
    existing = db.scalars(
        select(Product).where(Product.name == name, Product.brand == brand)
    ).first()
    if existing is not None:
        return existing
    product = Product(name=name, brand=brand, active=True)
    db.add(product)
    db.flush()
    return product


def _get_or_create_sku(
    db: Session,
    *,
    article: str,
    product_name: str,
    category: Category | None,
) -> Sku:
    sku = db.scalars(select(Sku).where(Sku.article == article)).first()
    brand = resolve_brand(sku.brand if sku else None, product_name, article)
    product = _get_or_create_product(db, product_name, brand)
    if sku is None:
        sku = Sku(
            article=article,
            name=product_name,
            product_id=product.id,
            category_id=category.id if category else None,
            brand=brand,
            unit="pcs",
            active=True,
        )
        db.add(sku)
        db.flush()
    else:
        sku.name = product_name
        sku.product_id = product.id
        sku.category_id = category.id if category else sku.category_id
        sku.brand = brand
        sku.unit = sku.unit or "pcs"
        sku.active = True
    if category is not None:
        db.execute(
            update(SalesFact).where(SalesFact.sku_id == sku.id).values(category_id=category.id)
        )
        db.execute(
            update(ReserveRow).where(ReserveRow.sku_id == sku.id).values(category_id=category.id)
        )
    return sku


def _has_sku_references(db: Session, sku_id: str) -> bool:
    reference_queries = (
        select(func.count()).select_from(SkuCost).where(SkuCost.sku_id == sku_id),
        select(func.count()).select_from(StockSnapshot).where(StockSnapshot.sku_id == sku_id),
        select(func.count()).select_from(SalesFact).where(SalesFact.sku_id == sku_id),
        select(func.count()).select_from(InboundDelivery).where(InboundDelivery.sku_id == sku_id),
        select(func.count()).select_from(ReserveRow).where(ReserveRow.sku_id == sku_id),
    )
    return any((db.scalar(query) or 0) > 0 for query in reference_queries)


def _delete_bad_category_skus(db: Session, category_labels: set[str]) -> int:
    deleted = 0
    if not category_labels:
        return deleted
    candidates = db.scalars(
        select(Sku).where(Sku.article.in_(category_labels), Sku.article == Sku.name)
    ).all()
    for sku in candidates:
        if _has_sku_references(db, sku.id):
            continue
        db.delete(sku)
        deleted += 1
    return deleted


def _append_status(upload_file: UploadFile, summary: AssortmentImportSummary) -> None:
    history = list(upload_file.batch.status_payload.get("history", []))
    history.append(
        {
            "status": "applied",
            "at": utc_now().isoformat(),
            "message": "Ассортимент импортирован как иерархия категорий по структуре XLSX",
            "error": None,
        }
    )
    upload_file.batch.status_payload = {
        **upload_file.batch.status_payload,
        "history": history[-25:],
        "current_status_at": utc_now().isoformat(),
        "last_error": None,
        "message": "Ассортимент связал артикулы с категориями по общей оси article/SKU",
        "assortment_import": {
            "category_rows": summary.category_rows,
            "product_rows": summary.product_rows,
            "linked_sku_rows": summary.linked_sku_rows,
            "deleted_bad_sku_rows": summary.deleted_bad_sku_rows,
            "skipped_rows": summary.skipped_rows,
        },
    }


def import_assortment_from_upload(
    db: Session,
    storage: ObjectStorage,
    upload_file_id: str,
    *,
    commit: bool = True,
) -> AssortmentImportSummary:
    upload_file = db.get(UploadFile, upload_file_id)
    if upload_file is None:
        raise ValueError(f"Upload file {upload_file_id} not found")

    payload = storage.load_upload_bytes(upload_file.storage_key)
    workbook = load_workbook(BytesIO(payload), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row, article_col, product_col = _find_header_columns(worksheet)

    stack: dict[int, Category] = {}
    category_labels: set[str] = set()
    category_rows = 0
    product_rows = 0
    linked_sku_rows = 0
    skipped_rows = 0

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        article_cell = worksheet.cell(row_idx, article_col)
        article = _cell_text(article_cell)
        product_name = _cell_text(worksheet.cell(row_idx, product_col))
        if not article:
            skipped_rows += 1
            continue

        outline_level = int(worksheet.row_dimensions[row_idx].outlineLevel or 0)
        if product_name:
            parent_candidates = [
                (level, category) for level, category in stack.items() if level < outline_level
            ]
            category = (
                max(parent_candidates, key=lambda item: item[0])[1] if parent_candidates else None
            )
            sku = _get_or_create_sku(
                db,
                article=article,
                product_name=product_name,
                category=category,
            )
            linked_sku_rows += 1 if sku.id else 0
            product_rows += 1
            continue

        parent_candidates = [
            (level, category) for level, category in stack.items() if level < outline_level
        ]
        parent = max(parent_candidates, key=lambda item: item[0])[1] if parent_candidates else None
        parent_path = parent.path if parent else ""
        path = f"{parent_path} / {article}" if parent_path else article
        category = _get_or_create_category(
            db,
            name=article,
            parent=parent,
            level=outline_level + 1,
            path=path,
        )
        stack = {level: item for level, item in stack.items() if level < outline_level}
        stack[outline_level] = category
        category_labels.add(article)
        category_rows += 1

    deleted_bad_sku_rows = _delete_bad_category_skus(db, category_labels)

    db.execute(delete(UploadedRowIssue).where(UploadedRowIssue.batch_id == upload_file.batch_id))
    db.execute(
        delete(QualityIssue).where(
            QualityIssue.batch_id == upload_file.batch_id,
            QualityIssue.file_id == upload_file.id,
        )
    )

    upload_file.source_type = "category_structure"
    upload_file.batch.source_type = "category_structure"
    upload_file.batch.detected_source_type = "category_structure"
    upload_file.batch.total_rows = max(worksheet.max_row - header_row, 0)
    upload_file.batch.valid_rows = product_rows
    upload_file.batch.applied_rows = product_rows
    upload_file.batch.failed_rows = 0
    upload_file.batch.warning_count = 0
    upload_file.batch.issue_count = 0
    upload_file.batch.status = "applied"
    upload_file.batch.mapping_payload = {
        **upload_file.batch.mapping_payload,
        "active_mapping": {
            worksheet.cell(header_row, article_col).value: "sku_code",
            worksheet.cell(header_row, product_col).value: "product_name",
        },
        "required_fields": ["sku_code", "product_name"],
        "assortment_adapter": {
            "name": "excel_outline_hierarchy_v1",
            "join_key": "article",
            "category_rows": category_rows,
            "product_rows": product_rows,
            "linked_sku_rows": linked_sku_rows,
            "deleted_bad_sku_rows": deleted_bad_sku_rows,
        },
        "source_detection": {
            **dict(upload_file.batch.mapping_payload.get("source_detection") or {}),
            "requires_confirmation": False,
            "confirmed": True,
            "detected_source_type": "category_structure",
            "selected_source_type": "category_structure",
        },
    }
    upload_file.batch.validation_payload = {
        "validated_at": utc_now().isoformat(),
        "valid_rows": product_rows,
        "failed_rows": 0,
        "warning_count": 0,
        "has_blocking_issues": False,
        "issue_counts": {"info": 0, "warning": 0, "error": 0, "critical": 0, "total": 0},
        "source_type": "category_structure",
        "adapter": "excel_outline_hierarchy_v1",
        "join_key": "article",
        "category_rows": category_rows,
        "product_rows": product_rows,
        "linked_sku_rows": linked_sku_rows,
        "deleted_bad_sku_rows": deleted_bad_sku_rows,
        "skipped_rows": skipped_rows,
    }

    summary = AssortmentImportSummary(
        upload_file_id=upload_file.id,
        file_name=upload_file.file_name,
        raw_rows=max(worksheet.max_row - header_row, 0),
        category_rows=category_rows,
        product_rows=product_rows,
        linked_sku_rows=linked_sku_rows,
        deleted_bad_sku_rows=deleted_bad_sku_rows,
        skipped_rows=skipped_rows,
    )
    _append_status(upload_file, summary)
    if commit:
        db.commit()
    return summary
