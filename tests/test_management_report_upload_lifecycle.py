from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from apps.api.app.db.models import ManagementReportImport


def _create_uploadable_management_report(path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    product_groups = workbook.create_sheet("Товарная группа")
    product_groups.append(["Товарная группа", "2024 г.", None, "2025 г.", None, None])
    product_groups.append(
        [None, "Выручка", "Рентабельность, %", "Выручка", "Рентабельность, %", None]
    )
    product_groups.append(["14. Толкатели и амортизаторы", 500, 57.0, 700, 65.03, 0.4])
    product_groups.append(["Итого", 500, 57.0, 700, 65.03, 0.4])

    departments = workbook.create_sheet("Подразделение")
    departments.append(["Подразделение", "2024 г.", None, "2025 г.", None, None])
    departments.append([None, "Выручка", "Рентабельность, %", "Выручка", "Рентабельность, %"])
    departments.append(["0311 ОТДЕЛ ПРОДАЖ КОМПЛЕКТУЮЩИЕ МОСКВА", 1000, 37.9, 1200, 39.1])
    departments.append(["0336 ОТДЕЛ СЕТЕВЫХ ПРОДАЖ", 2000, 60.66, 2400, 59.85])

    networks = workbook.create_sheet("Сети - разбивка по сетям")
    networks.append(["Контрагент", "2024 г.", None, "2025 г.", None, "Доля распределения", None])
    networks.append(
        [None, "Выручка", "Рентабельность, %", "Выручка", "Рентабельность, %", 2024, 2025]
    )
    networks.append(["ЛЕМАНА ПРО", 317857500.24, 62.29, 304926556.32, 60.78, 0.696106, 0.755264])

    workbook.save(path)


def test_management_report_upload_auto_detects_applies_and_keeps_lineage(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_uploadable_management_report(workbook_path)

    with workbook_path.open("rb") as workbook_file:
        response = client.post(
            "/api/uploads/files",
            data={},
            files={
                "file": (
                    workbook_path.name,
                    workbook_file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    file_payload = payload["file"]
    file_id = file_payload["id"]

    assert file_payload["source_type"] == "management_report"
    assert file_payload["detected_source_type"] == "management_report"
    assert file_payload["status"] == "applied"
    assert file_payload["applied_rows"] > 0
    assert file_payload["failed_rows"] == 0
    assert file_payload["source_detection"]["confirmed"] is True
    assert file_payload["source_detection"]["requires_confirmation"] is False

    import_record = db_session.scalar(
        select(ManagementReportImport).where(ManagementReportImport.upload_file_id == file_id)
    )
    assert import_record is not None
    assert import_record.file_name == "2025.xlsx"
    assert import_record.report_year == 2025
    assert import_record.metric_count > 0
    assert import_record.metadata_payload["upload_file_id"] == file_id
