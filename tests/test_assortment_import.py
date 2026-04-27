from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from apps.api.app.db.models import Category, Client, SalesFact, Sku, SkuCost


def _write_assortment_workbook(path: Path, article: str = "CSA-10L") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист_1"
    worksheet["A1"] = "Артикул"
    worksheet["D1"] = "Номенклатура"

    rows = [
        (2, "Магамакс", None, 0),
        (3, "Интерьер", None, 1),
        (4, "ДЕКОР", None, 2),
        (5, "01. Вазы", None, 3),
        (6, "01.01. Вазы стеклянные", None, 4),
        (7, article, "Ваза стеклянная тестовая", 5),
    ]
    for row_idx, article_value, product_name, outline_level in rows:
        worksheet.cell(row_idx, 1).value = article_value
        worksheet.cell(row_idx, 4).value = product_name
        worksheet.row_dimensions[row_idx].outlineLevel = outline_level
    workbook.save(path)


def test_assortment_upload_auto_imports_outline_categories_and_links_sku(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "Ассортимент.xlsx"
    _write_assortment_workbook(workbook_path)
    db_session.add(
        Sku(
            article="01. Вазы",
            name="01. Вазы",
            brand="MAGAMAX",
            unit="pcs",
            active=True,
        )
    )
    db_session.commit()

    with workbook_path.open("rb") as file:
        response = client.post(
            "/api/uploads/files",
            files={
                "file": (
                    workbook_path.name,
                    file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    file_payload = payload["file"]
    assert file_payload["source_type"] == "category_structure"
    assert file_payload["status"] == "applied"
    assert file_payload["applied_rows"] == 1
    assert file_payload["failed_rows"] == 0
    assert file_payload["issue_counts"]["total"] == 0
    assert file_payload["source_detection"]["requires_confirmation"] is False
    assert file_payload["readiness"]["can_apply"] is False
    assert file_payload["readiness"]["can_validate"] is False

    leaf = db_session.scalar(
        select(Category).where(
            Category.path == "Магамакс / Интерьер / ДЕКОР / 01. Вазы / 01.01. Вазы стеклянные"
        )
    )
    assert leaf is not None
    sku = db_session.scalar(select(Sku).where(Sku.article == "CSA-10L"))
    assert sku is not None
    assert sku.name == "Ваза стеклянная тестовая"
    assert sku.category_id == leaf.id
    assert db_session.scalar(select(Sku).where(Sku.article == "01. Вазы")) is None


def test_assortment_import_updates_existing_cost_and_sales_sku_category(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "Ассортимент.xlsx"
    _write_assortment_workbook(workbook_path, article="LINK-001")
    sku = Sku(
        article="LINK-001",
        name="Товар из себестоимости",
        brand="MAGAMAX",
        unit="pcs",
        active=True,
    )
    client_model = Client(
        code="obi-test",
        name="OBI test",
        region="Moscow",
        client_group="DIY",
        network_type="DIY",
        is_active=True,
    )
    db_session.add_all([sku, client_model])
    db_session.flush()
    db_session.add(
        SkuCost(
            article=sku.article,
            product_name=sku.name,
            cost_rub=Decimal("12.50"),
            sku_id=sku.id,
        )
    )
    fact = SalesFact(
        client_id=client_model.id,
        sku_id=sku.id,
        category_id=None,
        period_month=date(2025, 3, 1),
        quantity=10,
        revenue_amount=100,
    )
    db_session.add(fact)
    db_session.commit()

    with workbook_path.open("rb") as file:
        response = client.post(
            "/api/uploads/files",
            files={
                "file": (
                    workbook_path.name,
                    file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    db_session.refresh(sku)
    db_session.refresh(fact)
    assert sku.category_id is not None
    assert fact.category_id == sku.category_id
    cost = db_session.scalar(select(SkuCost).where(SkuCost.article == "LINK-001"))
    assert cost is not None
    assert cost.sku_id == sku.id
