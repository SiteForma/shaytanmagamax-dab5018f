from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from apps.api.app.db.models import (
    AssistantMessage,
    AssistantSession,
    ManagementReportImport,
    ManagementReportMetric,
    ManagementReportRow,
    OrganizationUnit,
)
from apps.api.app.common.utils import utc_now
from apps.api.app.modules.reports.management_import import import_management_report_workbook


def _create_management_report(path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    departments = workbook.create_sheet("Подразделение")
    departments.append(["Подразделение", "2024 г.", None, "2025 г.", None, None])
    departments.append([None, "Выручка", "Рентабельность, %", "Выручка", "Рентабельность, %", None])
    departments.append(
        [
            "0311 ОТДЕЛ ПРОДАЖ КОМПЛЕКТУЮЩИЕ МОСКВА",
            580396664.32,
            37.9,
            324404323.37,
            39.1,
            -0.4410644593,
        ]
    )
    departments.append(
        [
            "0336 ОТДЕЛ СЕТЕВЫХ ПРОДАЖ",
            456622247.96,
            60.66,
            403735154.79,
            59.85,
            -0.1158224187,
        ]
    )
    departments.append(["Итого", 1037018912.28, 48.22, 728139478.16, 50.21, -0.2978532315])
    departments.append(["СПРОС", 270320198.03, None, 395087629.57, None, None])

    networks = workbook.create_sheet("Сети - разбивка по сетям")
    networks.append(["Контрагент", "2024 г.", None, "2025 г.", None, "Доля распределения", None])
    networks.append([None, "Выручка", "Рентабельность, %", "Выручка", "Рентабельность, %", 2024, 2025])
    networks.append(["ЛЕМАНА ПРО", 317857500.24, 62.29, 304926556.32, 60.78, 0.696106, 0.755264])

    product_groups = workbook.create_sheet("Товарная группа")
    product_groups.append(["Товарная группа", "2024 г.", None, "2025 г.", None, None])
    product_groups.append([None, "Выручка", "Рентабельность, %", "Выручка", "Рентабельность, %", None])
    product_groups.append(["01. Ручки мебельные", 1000, 48.0, 1500, 47.04, 0.5])
    product_groups.append(["07. Петли мебельные", 2000, 52.0, 1800, 60.69, -0.1])
    product_groups.append(["14. Толкатели и амортизаторы", 500, 57.0, 700, 65.03, 0.4])
    product_groups.append(["Итого", 3500, 50.0, 4000, 55.0, 0.14])

    pdz = workbook.create_sheet("ПДЗ")
    pdz.append(["Контрагент", "2024 г.", None, "2025 г.", None, "Сокращение ПДЗ"])
    pdz.append([None, "Всего", "ПДЗ", "Всего", "ПДЗ", None])
    pdz.append(
        [
            "0336 ОТДЕЛ СЕТЕВЫХ ПРОДАЖ",
            178052353.05,
            41319718.84,
            92868471.24,
            2879691.96,
            -0.9303070776,
        ]
    )

    demand = workbook.create_sheet("СПРОС (недопоставки)")
    demand.append(["2024 г.", "2025 г."])
    demand.append([270320198.03, 395087629.57])

    workbook.save(path)


def test_management_report_import_persists_rows_metrics_and_departments(
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)

    summary = import_management_report_workbook(db_session, workbook_path, report_year=2025)

    assert summary.sheet_count == 5
    assert summary.raw_row_count == 20
    assert summary.metric_count > 0
    assert summary.organization_unit_count == 2

    import_record = db_session.scalar(
        select(ManagementReportImport).where(ManagementReportImport.id == summary.import_id)
    )
    assert import_record is not None
    assert import_record.metadata_payload["parser_version"] == "management_report_v1"

    department = db_session.scalar(
        select(OrganizationUnit).where(
            OrganizationUnit.unit_type == "department",
            OrganizationUnit.code == "0311",
        )
    )
    assert department is not None
    assert department.name == "ОТДЕЛ ПРОДАЖ КОМПЛЕКТУЮЩИЕ МОСКВА"

    revenue_2025 = db_session.scalar(
        select(ManagementReportMetric.metric_value).where(
            ManagementReportMetric.sheet_name == "Подразделение",
            ManagementReportMetric.dimension_code == "0311",
            ManagementReportMetric.metric_name == "revenue",
            ManagementReportMetric.metric_year == 2025,
        )
    )
    assert float(revenue_2025) == 324404323.37

    demand_metric = db_session.scalar(
        select(ManagementReportMetric.metric_value).where(
            ManagementReportMetric.dimension_type == "demand_shortage",
            ManagementReportMetric.metric_name == "demand_shortage",
            ManagementReportMetric.metric_year == 2025,
        )
    )
    assert float(demand_metric) == 395087629.57


def test_management_report_import_is_idempotent_for_same_file(
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)

    first = import_management_report_workbook(db_session, workbook_path, report_year=2025)
    second = import_management_report_workbook(db_session, workbook_path, report_year=2025)

    assert second.import_id == first.import_id
    assert db_session.scalar(select(func.count()).select_from(ManagementReportImport)) == 1
    assert db_session.scalar(select(func.count()).select_from(ManagementReportRow)) == first.raw_row_count
    assert (
        db_session.scalar(select(func.count()).select_from(ManagementReportMetric))
        == first.metric_count
    )
    assert db_session.scalar(select(func.count()).select_from(OrganizationUnit)) == 2


def test_management_report_api_returns_summary_and_metrics(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    summary_response = client.get("/api/reports/management/summary")
    assert summary_response.status_code == 200
    summary = summary_response.json()
    assert summary["latestImport"]["fileName"] == "2025.xlsx"
    assert summary["latestImport"]["metricCount"] > 0
    assert len(summary["organizationUnits"]) == 2

    metrics_response = client.get(
        "/api/reports/management/metrics",
        params={
            "sheetName": "Подразделение",
            "metricName": "revenue",
            "metricYear": 2025,
        },
    )
    assert metrics_response.status_code == 200
    metrics = metrics_response.json()
    assert any(item["dimensionCode"] == "0311" for item in metrics)


def test_assistant_answers_from_management_report(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    response = client.post(
        "/api/assistant/query",
        json={"text": "Покажи управленческий отчет 2025 по подразделениям и выручке"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["intent"] == "management_report_summary"
    assert payload["status"] == "completed"
    assert any(ref["sourceType"] == "management_report" for ref in payload["sourceRefs"])
    assert any(tool["toolName"] == "get_management_report" for tool in payload["toolCalls"])


def test_assistant_answers_product_group_profitability_short_query(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    response = client.post(
        "/api/assistant/query",
        json={"text": "какая ТГ самая выгодная в 2025"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["intent"] == "management_report_summary"
    assert payload["status"] == "completed"
    assert payload["title"] == "Самая выгодная ТГ в 2025"
    assert "Толкатели и амортизаторы" in payload["summary"]
    assert payload["sections"][1]["rows"][0]["dimensionName"] == "14. Толкатели и амортизаторы"
    assert any(ref["sourceType"] == "management_report" for ref in payload["sourceRefs"])


def test_assistant_answers_product_group_earnings_natural_query(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    response = client.post(
        "/api/assistant/query",
        json={"text": "на каком товара мы заработали в 2025 больше всего?"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["intent"] == "management_report_summary"
    assert payload["status"] == "completed"
    assert "Петли мебельные" in payload["summary"]
    assert payload["sections"][1]["rows"][0]["dimensionName"] == "07. Петли мебельные"
    assert "выручка × рентабельность" in payload["sections"][2]["items"][1]


def test_assistant_followup_expands_previous_management_report_answer(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    session_response = client.post("/api/assistant/sessions", json={"title": "Отчёт 2025"})
    assert session_response.status_code == 200
    session = session_response.json()

    first_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "какая ТГ самая выгодная в 2025"},
    )
    assert first_response.status_code == 200
    assert first_response.json()["response"]["intent"] == "management_report_summary"

    followup_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "дай развернутые данные"},
    )

    assert followup_response.status_code == 200
    payload = followup_response.json()["response"]
    assert payload["intent"] == "management_report_summary"
    assert payload["status"] == "completed"
    assert any(tool["toolName"] == "get_management_report" for tool in payload["toolCalls"])
    assert payload["sections"][1]["rows"][0]["dimensionName"] == "14. Толкатели и амортизаторы"


def test_assistant_followup_can_ask_second_place_from_previous_report_answer(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    session_response = client.post("/api/assistant/sessions", json={"title": "Отчёт 2025"})
    assert session_response.status_code == 200
    session = session_response.json()

    first_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "на каком товаре мы заработали в 2025 больше всего?"},
    )
    assert first_response.status_code == 200
    assert first_response.json()["response"]["intent"] == "management_report_summary"

    followup_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "а второе место?"},
    )

    assert followup_response.status_code == 200
    payload = followup_response.json()["response"]
    assert payload["intent"] == "management_report_summary"
    assert payload["status"] == "completed"
    assert "2-е место" in payload["summary"]
    assert "Ручки мебельные" in payload["summary"]
    assert any(tool["toolName"] == "get_management_report" for tool in payload["toolCalls"])


def test_assistant_followup_recovers_context_after_failed_short_followup(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    session_response = client.post("/api/assistant/sessions", json={"title": "Отчёт 2025"})
    assert session_response.status_code == 200
    session = session_response.json()

    first_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "на каком товаре мы заработали в 2025 больше всего?"},
    )
    assert first_response.status_code == 200

    # Simulate an already existing failed turn in a real user session.
    failed_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "Кто написал Евгения Онегина?"},
    )
    assert failed_response.status_code == 200
    assert failed_response.json()["response"]["intent"] == "unsupported_or_ambiguous"

    followup_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "а второе место?"},
    )

    assert followup_response.status_code == 200
    payload = followup_response.json()["response"]
    assert payload["intent"] == "management_report_summary"
    assert "2-е место" in payload["summary"]
    assert "Ручки мебельные" in payload["summary"]


def test_assistant_repair_followup_answers_previous_failed_short_question(
    client: TestClient,
    db_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "2025.xlsx"
    _create_management_report(workbook_path)
    import_management_report_workbook(db_session, workbook_path, report_year=2025)

    session_response = client.post("/api/assistant/sessions", json={"title": "Отчёт 2025"})
    assert session_response.status_code == 200
    session = session_response.json()

    first_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "на каком товаре мы заработали в 2025 больше всего?"},
    )
    assert first_response.status_code == 200

    stored_session = db_session.get(AssistantSession, session["id"])
    assert stored_session is not None
    db_session.add(
        AssistantMessage(
            session_id=session["id"],
            created_by_id="user_admin",
            role="user",
            message_text="а второе место?",
            context_payload={},
            generated_at=utc_now(),
        )
    )
    db_session.add(
        AssistantMessage(
            session_id=session["id"],
            created_by_id="user_admin",
            role="assistant",
            message_text="Отказ",
            intent="unsupported_or_ambiguous",
            status="unsupported",
            provider="deterministic",
            confidence=0.95,
            trace_id="trace_old_failure",
            context_payload={},
            response_payload={
                "intent": "unsupported_or_ambiguous",
                "status": "unsupported",
                "summary": "Отказ",
                "toolCalls": [],
                "sourceRefs": [],
            },
            generated_at=utc_now(),
        )
    )
    db_session.add(
        AssistantMessage(
            session_id=session["id"],
            created_by_id="user_admin",
            role="user",
            message_text="ну а теперь ответишь?",
            context_payload={},
            generated_at=utc_now(),
        )
    )
    db_session.add(
        AssistantMessage(
            session_id=session["id"],
            created_by_id="user_admin",
            role="assistant",
            message_text="Подключён управленческий отчёт 2025.xlsx",
            intent="management_report_summary",
            status="completed",
            provider="deterministic",
            confidence=0.7,
            trace_id="trace_old_generic_answer",
            context_payload={},
            response_payload={
                "intent": "management_report_summary",
                "status": "completed",
                "summary": "Подключён управленческий отчёт 2025.xlsx",
                "toolCalls": [
                    {
                        "toolName": "get_management_report",
                        "arguments": {"question": "ну а теперь ответишь?"},
                    }
                ],
                "sourceRefs": [],
            },
            tool_calls=[
                {
                    "toolName": "get_management_report",
                    "arguments": {"question": "ну а теперь ответишь?"},
                }
            ],
            generated_at=utc_now(),
        )
    )
    stored_session.message_count += 4
    db_session.commit()

    repair_response = client.post(
        f"/api/assistant/sessions/{session['id']}/messages",
        json={"text": "ну а теперь ответишь?"},
    )

    assert repair_response.status_code == 200
    payload = repair_response.json()["response"]
    assert payload["intent"] == "management_report_summary"
    assert "2-е место" in payload["summary"]
    assert "Ручки мебельные" in payload["summary"]
